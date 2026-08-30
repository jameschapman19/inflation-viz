"""Fetcher for the current live CPI/CPIH basket weights.

Unlike the CDID timeseries series, weights aren't published behind a stable
JSON API — they're a downloadable workbook on the dataset page named in
`sources.yaml`'s `reference_tables.basket_weights` entry, refreshed each
Jan/Feb. This module resolves the current download link from that page,
then parses the workbook's COICOP division weights.

Caveat for whoever runs this next: this dev sandbox has ons.gov.uk blocked
by network egress policy, so the link-resolution regex and the sheet layout
below could not be checked against the live page or workbook. Both are
written against a best-effort understanding of ONS's publishing conventions
and covered by fixture-based tests (tests/fixtures/ons_weights_page.html,
tests/fixtures/ons_weights_workbook.xlsx). Before relying on this in
production, run it once against the live page and adjust
`_DOWNLOAD_LINK_PATTERN` / `_WEIGHTS_SHEET_NAME` / `_parse_weights_sheet` to
match what's actually there.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO

import openpyxl
import requests

from inflation_viz.config import ReferenceTableSource

_DOWNLOAD_LINK_PATTERN = re.compile(
    r'href="(?P<href>[^"]*(?:weight|Table2)[^"]*\.(?:xlsx|xls|ods))"', re.IGNORECASE
)
_WEIGHTS_SHEET_NAME = "Weights"
_COICOP_CODE_RE = re.compile(r"^(0[1-9]|1[0-2])\b")


@dataclass(frozen=True, slots=True)
class DivisionWeight:
    coicop: str
    division_name: str
    weight_per_mille: float


def resolve_weights_download_url(source: ReferenceTableSource, session: requests.Session) -> str:
    response = session.get(source.source_url, timeout=30)
    response.raise_for_status()
    match = _DOWNLOAD_LINK_PATTERN.search(response.text)
    if match is None:
        raise ValueError(
            f"Could not find a weights workbook download link on {source.source_url}. "
            "The ONS page layout may have changed — update _DOWNLOAD_LINK_PATTERN."
        )
    href = match.group("href")
    if href.startswith("http"):
        return href
    return f"https://www.ons.gov.uk{href}"


def parse_weights_workbook(content: bytes) -> list[DivisionWeight]:
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if _WEIGHTS_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Expected a '{_WEIGHTS_SHEET_NAME}' sheet, found {workbook.sheetnames}. "
            "Update _WEIGHTS_SHEET_NAME to match the live workbook."
        )
    sheet = workbook[_WEIGHTS_SHEET_NAME]

    weights: list[DivisionWeight] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code_cell = str(row[0]).strip()
        match = _COICOP_CODE_RE.match(code_cell)
        if match is None:
            continue
        coicop = match.group(1)
        division_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        weight_raw = row[2] if len(row) > 2 else None
        if weight_raw is None:
            continue
        weights.append(
            DivisionWeight(
                coicop=coicop,
                division_name=division_name,
                weight_per_mille=float(str(weight_raw)),
            )
        )
    return weights


def fetch_weights(
    source: ReferenceTableSource, session: requests.Session | None = None
) -> list[DivisionWeight]:
    owns_session = session is None
    session = session or requests.Session()
    try:
        url = resolve_weights_download_url(source, session)
        response = session.get(url, timeout=30)
        response.raise_for_status()
        return parse_weights_workbook(response.content)
    finally:
        if owns_session:
            session.close()


def fetched_at_now() -> datetime:
    return datetime.now(UTC)
